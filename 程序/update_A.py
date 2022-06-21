import os
import time
import datetime
import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

alignment = Alignment(horizontal='center', vertical='center')

def update_table():
    global a_id_cards_list, a_id_cards_dic, a_sht1_max_row, a_sht2_max_row
    for i in range(len(d_id_cards_list)):
        print("处理第%s条"%i)
        # 如果D表中该用户的身份证可以在A表身份证列查到
        if d_id_cards_list[i] in a_id_cards_dic.keys():
            a_index = a_id_cards_dic[d_id_cards_list[i]]
            # 判断表D的代发账号是否在表B的三列中存在
            b_index = -1
            if d_issue_account_list[i] in b_account_dic.keys():
                b_index = b_account_dic[d_issue_account_list[i]]
            if d_issue_account_list[i] in b_old_account_dic.keys():
                b_index = b_old_account_dic[d_issue_account_list[i]]
            if d_issue_account_list[i] in b_card_dic.keys():
                b_index = b_card_dic[d_issue_account_list[i]]
            # 如果D表的账号在b表三列中存在
            if b_index >= 0:
                # 如果A表账户在B表三个值中存在
                if a_issue_account_list[a_index] == b_account_list[b_index] or a_issue_account_list[a_index] == b_card_list[b_index] or a_issue_account_list[a_index] == b_old_account_list[b_index]:
                    # D表M列，“已存在，无需装表”
                    d_ws['M' + str(i+2)] = "已存在，无需装表"
                    d_ws['M' + str(i+2)].alignment = alignment
                # 如果A表的账户在B表中不存在
                else:
                    # 如果B表的身份证为空
                    if not b_id_cards_list[b_index]:
                        d_ws['L' + str(i+2)] = "空值"
                        d_ws['L' + str(i+2)].alignment = alignment
                        d_ws['M' + str(i+2)] = "数据异常，不予装表"
                        d_ws['M' + str(i+2)].alignment = alignment
                    # B表的身份证不为空
                    else:
                        # 判断D表的身份证和B表是否一致
                        if b_id_cards_list[b_index] == d_id_cards_list[i]:
                            # 判断B表的状态是否为"关户"
                            if b_status_list[b_index] == "关户":
                                a_sht2_max_row = a_sheet2.max_row
                                # 复制A表sheet1的整行数据至A表sheet2，
                                a_item_tuple = a_sheet1[a_index + 2]
                                for j in range(len(a_item_tuple)):
                                    cell = a_item_tuple[j].value
                                    letter = chr(j + 65)
                                    a_sheet2[letter + str(a_sht2_max_row+1)] = cell
                                    a_sheet2[letter + str(a_sht2_max_row+1)].alignment = alignment
                                # 格式化A表sheet2的装表日期
                                a_sheet2['H' + str(a_sht2_max_row+1)].number_format = 'yyyy/m/d'
                                a_sheet2['H' + str(a_sht2_max_row+1)].alignment = alignment
                                # A表sheet2赋值I、J列为删除日期和B表对应状态
                                a_sheet2['I' + str(a_sht2_max_row+1)] = datetime.date(year,month,day)
                                a_sheet2['I' + str(a_sht2_max_row+1)].number_format = 'yyyy/m/d'
                                a_sheet2['I' + str(a_sht2_max_row+1)].alignment = alignment
                                a_sheet2['J' + str(a_sht2_max_row+1)] = b_status_list[b_index]
                                a_sheet2['J' + str(a_sht2_max_row+1)].alignment = alignment
                                # 删除A表sheet1该行数据
                                a_sheet1.delete_rows(a_index + 2)
                                # 更新最大行数值
                                a_sht1_max_row = a_sheet1.max_row
                                # D表数据写入A表sheet1最后一行
                                # A表B列代发单位 == D表J列代发单位
                                a_sheet1['B' + str(a_sht1_max_row+1)] = d_ws['J' + str(i+2)].value
                                a_sheet1['B' + str(a_sht1_max_row+1)].alignment = alignment
                                # A表C列一卡通身份证号 == D表G列代发身份证号
                                a_sheet1['C' + str(a_sht1_max_row+1)] = d_ws['G' + str(i+2)].value
                                a_sheet1['C' + str(a_sht1_max_row+1)].alignment = alignment
                                # A表D列一卡通账户 == D表B列代发账户
                                a_sheet1['D' + str(a_sht1_max_row+1)] = d_ws['B' + str(i+2)].value
                                a_sheet1['D' + str(a_sht1_max_row+1)].alignment = alignment
                                # A表E列一卡通户名 == D表C列代发户名
                                a_sheet1['E' + str(a_sht1_max_row+1)] = d_ws['C' + str(i+2)].value
                                a_sheet1['E' + str(a_sht1_max_row+1)].alignment = alignment
                                # A表F列金额 == D表E列代发金额
                                a_sheet1['F' + str(a_sht1_max_row+1)] = d_ws['E' + str(i+2)].value
                                a_sheet1['F' + str(a_sht1_max_row+1)].alignment = alignment
                                # A表G列一卡通补贴项目 == D表K列补贴项目
                                a_sheet1['G' + str(a_sht1_max_row+1)] = d_ws['K' + str(i+2)].value
                                a_sheet1['G' + str(a_sht1_max_row+1)].alignment = alignment
                                # A表sheet1赋值H列装表日期
                                a_sheet1['H' + str(a_sht1_max_row+1)] = datetime.date(year,month,day)
                                a_sheet1['H' + str(a_sht1_max_row+1)].number_format = 'yyyy/m/d'
                                a_sheet1['H' + str(a_sht1_max_row+1)].alignment = alignment
                                # D表M列写入装表结果
                                d_ws['L' + str(i+2)] = b_id_cards_list[b_index]
                                d_ws['L' + str(i+2)].alignment = alignment
                                d_ws['M' + str(i+2)] = '新装表'
                                d_ws['M' + str(i+2)].alignment = alignment
                                # 更新a表身份证列表和字典（删除一个，并增加一个）
                                del a_id_cards_list[a_index]
                                a_id_cards_list.append(a_sheet1['C' + str(a_sht1_max_row+1)])
                                a_id_cards_dic = dict(zip(a_id_cards_list, [x for x in range(len(a_id_cards_list))]))
                            # B表状态不是"关户"
                            else:
                                # 如果D表的账号是6230开头，或者，D表的账号8开头且和B表账号一致，则更新A表
                                if d_issue_account_list[i].startswith('6230') or (d_issue_account_list[i].startswith('8') and d_issue_account_list[i] == b_account_list[b_index]):
                                    a_sht2_max_row = a_sheet2.max_row
                                    # 复制A表sheet1的整行数据至A表sheet2，A表sheet2赋值I、J列为删除日期和B表对应状态
                                    a_item_tuple = a_sheet1[a_index + 2]
                                    for j in range(len(a_item_tuple)):
                                        cell = a_item_tuple[j].value
                                        letter = chr(j + 65)
                                        a_sheet2[letter + str(a_sht2_max_row+1)] = cell
                                        a_sheet2[letter + str(a_sht2_max_row+1)].alignment = alignment
                                    # 格式化A表sheet2的装表日期
                                    a_sheet2['H' + str(a_sht2_max_row+1)].number_format = 'yyyy/m/d'
                                    a_sheet2['H' + str(a_sht2_max_row+1)].alignment = alignment
                                    a_sheet2['I' + str(a_sht2_max_row+1)] = datetime.date(year,month,day)
                                    a_sheet2['I' + str(a_sht2_max_row+1)].number_format = 'yyyy/m/d'
                                    a_sheet2['I' + str(a_sht2_max_row+1)].alignment = alignment
                                    a_sheet2['J' + str(a_sht2_max_row+1)] = b_status_list[b_index]
                                    a_sheet2['J' + str(a_sht2_max_row+1)].alignment = alignment
                                    # 删除A表sheet1该行数据
                                    a_sheet1.delete_rows(a_index + 2)
                                    a_sht1_max_row = a_sheet1.max_row
                                    # D表数据写入A表sheet1最后一行
                                    # A表B列代发单位 == D表J列代发单位
                                    a_sheet1['B' + str(a_sht1_max_row+1)] = d_ws['J' + str(i+2)].value
                                    a_sheet1['B' + str(a_sht1_max_row+1)].alignment = alignment
                                    # A表C列一卡通身份证号 == D表G列代发身份证号
                                    a_sheet1['C' + str(a_sht1_max_row+1)] = d_ws['G' + str(i+2)].value
                                    a_sheet1['C' + str(a_sht1_max_row+1)].alignment = alignment
                                    # A表D列一卡通账户 == D表B列代发账户
                                    a_sheet1['D' + str(a_sht1_max_row+1)] = d_ws['B' + str(i+2)].value
                                    a_sheet1['D' + str(a_sht1_max_row+1)].alignment = alignment
                                    # A表E列一卡通户名 == D表C列代发户名
                                    a_sheet1['E' + str(a_sht1_max_row+1)] = d_ws['C' + str(i+2)].value
                                    a_sheet1['E' + str(a_sht1_max_row+1)].alignment = alignment
                                    # A表F列金额 == D表E列代发金额
                                    a_sheet1['F' + str(a_sht1_max_row+1)] = d_ws['E' + str(i+2)].value
                                    a_sheet1['F' + str(a_sht1_max_row+1)].alignment = alignment
                                    # A表G列一卡通补贴项目 == D表K列补贴项目
                                    a_sheet1['G' + str(a_sht1_max_row+1)] = d_ws['K' + str(i+2)].value
                                    a_sheet1['G' + str(a_sht1_max_row+1)].alignment = alignment
                                    # A表sheet1赋值H列装表日期
                                    a_sheet1['H' + str(a_sht1_max_row+1)] = datetime.date(year,month,day)
                                    a_sheet1['H' + str(a_sht1_max_row+1)].number_format = 'yyyy/m/d'
                                    a_sheet1['H' + str(a_sht1_max_row+1)].alignment = alignment
                                    # D表M列写入装表结果
                                    d_ws['L' + str(i+2)] = b_id_cards_list[b_index]
                                    d_ws['L' + str(i+2)].alignment = alignment
                                    d_ws['M' + str(i+2)] = '新装表'
                                    d_ws['M' + str(i+2)].alignment = alignment
                                    # 更新a表身份证列表和字典（删除一个，并增加一个）
                                    del a_id_cards_list[a_index]
                                    a_id_cards_list.append(a_sheet1['C' + str(a_sht1_max_row+1)])
                                    a_id_cards_dic = dict(zip(a_id_cards_list, [x for x in range(len(a_id_cards_list))]))                           
                        # D表身份证和B表的不一致
                        else:
                            d_ws['L' + str(i+2)] = b_id_cards_list[b_index]
                            d_ws['L' + str(i+2)].alignment = alignment
                            d_ws['M' + str(i+2)] = "数据异常，不予装表"
                            d_ws['M' + str(i+2)].alignment = alignment
            # D表账号在B表不存在
            else:
                d_ws['M' + str(i+2)] = "暂未提取到数据，不予装表"
                d_ws['M' + str(i+2)].alignment = alignment
        # D表身份证在A表查不到
        else:
            # 判断D表的代发账号是否在B表的三列（账号、卡号、旧帐号）数据中存在
            b_index = -1
            if d_issue_account_list[i] in b_account_dic.keys():
                b_index = b_account_dic[d_issue_account_list[i]]
            if d_issue_account_list[i] in b_old_account_dic.keys():
                b_index = b_old_account_dic[d_issue_account_list[i]]
            if d_issue_account_list[i] in b_card_dic.keys():
                b_index = b_card_dic[d_issue_account_list[i]]  
            # 如果D表的代发账号在B表的三列存在
            if b_index >= 0:
                if not b_id_cards_list[b_index]:
                    d_ws['L' + str(i+2)] = "空值"
                    d_ws['L' + str(i+2)].alignment = alignment
                    d_ws['M' + str(i+2)] = "数据异常，不予装表"
                    d_ws['M' + str(i+2)].alignment = alignment
                else:
                    if b_id_cards_list[b_index] == d_id_cards_list[i]:                                                
                        a_sht1_max_row = a_sheet1.max_row
                        # D表数据写入A表sheet1最后一行
                        # A表B列代发单位 == D表J列代发单位
                        a_sheet1['B' + str(a_sht1_max_row+1)] = d_ws['J' + str(i+2)].value
                        a_sheet1['B' + str(a_sht1_max_row+1)].alignment = alignment
                        # A表C列一卡通身份证号 == D表G列代发身份证号
                        a_sheet1['C' + str(a_sht1_max_row+1)] = d_ws['G' + str(i+2)].value
                        a_sheet1['C' + str(a_sht1_max_row+1)].alignment = alignment
                        # A表D列一卡通账户 == D表B列代发账户
                        a_sheet1['D' + str(a_sht1_max_row+1)] = d_ws['B' + str(i+2)].value
                        a_sheet1['D' + str(a_sht1_max_row+1)].alignment = alignment
                        # A表E列一卡通户名 == D表C列代发户名
                        a_sheet1['E' + str(a_sht1_max_row+1)] = d_ws['C' + str(i+2)].value
                        a_sheet1['E' + str(a_sht1_max_row+1)].alignment = alignment
                        # A表F列金额 == D表E列代发金额
                        a_sheet1['F' + str(a_sht1_max_row+1)] = d_ws['E' + str(i+2)].value
                        a_sheet1['F' + str(a_sht1_max_row+1)].alignment = alignment
                        # A表G列一卡通补贴项目 == D表K列补贴项目
                        a_sheet1['G' + str(a_sht1_max_row+1)] = d_ws['K' + str(i+2)].value
                        a_sheet1['G' + str(a_sht1_max_row+1)].alignment = alignment
                        # A表sheet1赋值H列装表日期
                        a_sheet1['H' + str(a_sht1_max_row+1)] = datetime.date(year,month,day)
                        a_sheet1['H' + str(a_sht1_max_row+1)].number_format = 'yyyy/m/d'
                        a_sheet1['H' + str(a_sht1_max_row+1)].alignment = alignment
                        # D表M列写入装表结果
                        d_ws['L' + str(i+2)] = b_id_cards_list[b_index]
                        d_ws['L' + str(i+2)].alignment = alignment
                        d_ws['M' + str(i+2)] = '新装表'
                        d_ws['M' + str(i+2)].alignment = alignment
                        # 更新a表身份证列表和字典（增加一个）
                        a_id_cards_list.append(a_sheet1['C' + str(a_sht1_max_row+1)])
                        a_id_cards_dic = dict(zip(a_id_cards_list, [x for x in range(len(a_id_cards_list))]))
                    else:
                        d_ws['L' + str(i+2)] = b_id_cards_list[b_index]
                        d_ws['L' + str(i+2)].alignment = alignment
                        d_ws['M' + str(i+2)] = "数据异常，不予装表"
                        d_ws['M' + str(i+2)].alignment = alignment
            # 如果D表的代发账号在B表的三列不存在
            else:
                # D表M列“暂未提取到数据，不予装表”
                d_ws['M' + str(i+2)] = "暂未提取到数据，不予装表"
                d_ws['M' + str(i+2)].alignment = alignment


if __name__ == '__main__':
    print("开始执行！")
    start = time.perf_counter()
    local = time.strftime("%Y%m%d%H%M%S", time.localtime())
    date = time.localtime()
    year, month, day = date.tm_year, date.tm_mon, date.tm_mday


    a_file_name = "表A-手工数据母表.xlsx"
    a_sht1_name = '母表'
    a_sht2_name = '已删除账号信息表'

    b_file_name = '表B-后台导出数据.xlsx'
    b_sht_name = 'Sheet1'

    d_file_name = "表D-结果数据.xlsx"
    d_sht_name = 'Sheet1'

    # 更新后的A表\D表名称
    a_new_file_name = '更新后的表A.xlsx'
    d_new_file_name = "更新后的表D.xlsx"

    base_dir = os.path.dirname(os.path.abspath(__file__))
    a_file_path = os.path.join(base_dir, a_file_name)
    b_file_path = os.path.join(base_dir, b_file_name)    
    d_file_path = os.path.join(base_dir, d_file_name)

    d_wb = load_workbook(d_file_path)
    d_ws = d_wb[d_sht_name]
    d_id_cards_tuple = d_ws['G'][1:]
    d_issue_account_tuple = d_ws['B'][1:]
    d_id_cards_list = []
    d_issue_account_list = []
    for i in d_id_cards_tuple:
        d_id_cards_list.append(i.value)
    for i in d_issue_account_tuple:
        d_issue_account_list.append(i.value)
    print("D表数据获取完成")

    a_table = load_workbook(a_file_path)
    a_sheet1 = a_table[a_sht1_name]
    a_sheet2 = a_table[a_sht2_name]

    a_id_cards_tuple = a_sheet1['C'][1:]
    a_issue_account_tuple = a_sheet1['D'][1:]
    a_id_cards_list = []
    a_issue_account_list = []
    for i in a_id_cards_tuple:
        a_id_cards_list.append(i.value)
    for i in a_issue_account_tuple:
        a_issue_account_list.append(i.value)

    a_idx = [x for x in range(len(a_id_cards_list))]
    a_id_cards_dic = dict(zip(a_id_cards_list, a_idx))

    a_sht1_max_row = a_sheet1.max_row
    a_sht2_max_row = a_sheet2.max_row
    print("A表数据获取完成")

    b_table = xlrd.open_workbook(b_file_path)
    b_sheet = b_table.sheet_by_name(b_sht_name)
    b_title_list = b_sheet.row_values(0)
    for i in b_title_list:
        if i == '账号':
            account_idx = b_title_list.index(i)
        if i == '卡号':
            card_idx = b_title_list.index(i)
        if i == '旧账号':
            old_account_idx = b_title_list.index(i)
        if i == '证件号码':
            id_cards_idx = b_title_list.index(i)
        if i == "当前状态":
            status_idx = b_title_list.index(i)
    b_account_list = b_sheet.col_values(account_idx)[1:]
    b_card_list = b_sheet.col_values(card_idx)[1:]
    b_old_account_list = b_sheet.col_values(old_account_idx)[1:]
    b_id_cards_list = b_sheet.col_values(id_cards_idx)[1:]
    b_status_list = b_sheet.col_values(status_idx)[1:]
    
    b_idx = [x for x in range(len(b_account_list))]
    b_account_dic = dict(zip(b_account_list, b_idx))
    b_card_dic = dict(zip(b_card_list, b_idx))
    b_old_account_dic = dict(zip(b_old_account_list, b_idx))
    print("B表数据获取完成")

    update_table()

    b_table.release_resources()
    del b_table
    # 保存更新后的A表和D表
    a_new_file_name = a_new_file_name.split('.')[0] + local + '.' + a_new_file_name.split('.')[1]
    a_new_file_path = os.path.join(base_dir, a_new_file_name)
    a_table.save(a_new_file_path)
    d_new_file_name = d_new_file_name.split('.')[0] + local + '.' + d_new_file_name.split('.')[1]
    d_new_file_path = os.path.join(base_dir, d_new_file_name)
    d_wb.save(d_new_file_path)

    end = time.perf_counter()
    print("执行成功！")
    print('更新后的A表保存至: %s'%a_new_file_path)
    print('更新后的D表保存至: %s'%d_new_file_path)
    print('Running time: %s Seconds' %(end-start))
