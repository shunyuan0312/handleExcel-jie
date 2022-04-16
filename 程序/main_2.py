import os
import time
import xlrd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

dic = {'1':'A','2':'B','3':'C','4':'D','5':'E','6':'F','7':'G','8':'H','9':'I','10':'J','11':'K','12':'L','13':'M','14':'N','15':'O','16':'P','17':'Q','18':'R','19':'S','20':'T','21':'U','22':'V','23':'W','24':'X','25':'Y','26':'Z'}
red_bg_fill = PatternFill(fill_type="solid",fgColor="ff0000")
orange_bg_fill = PatternFill(fill_type="solid",fgColor="ffa500")
yellow_bg_fill = PatternFill(fill_type="solid",fgColor="ffff00")

def down_to_fifteen(eighteen_card):
    return eighteen_card[0:6] + eighteen_card[8:17]

def set_row_bg(bg_fill, row, min_col=1, max_col=16):
    for rows in c_ws.iter_rows(min_row=row, max_row=row, min_col=min_col, max_col=max_col):
        for cell in rows:
            cell.fill = bg_fill

def update_tableC():    
    for i in range(len(c_id_cards_list)):
        print("处理第%s条"%i)
        # 如果C表中该用户的身份证可以在A表身份证列查到
        if c_id_cards_list[i] in a_id_cards_list:
            a_index = a_id_cards_list.index(c_id_cards_list[i])
            # 判断表c的代发账号是否在表b的三列中存在
            row_num = -1
            if c_issue_account_list[i] in b_account_list:
                row_num = b_account_list.index(c_issue_account_list[i])
            if c_issue_account_list[i] in b_card_list:
                row_num = b_card_list.index(c_issue_account_list[i])
            if c_issue_account_list[i] in b_old_account_list:
                row_num = b_old_account_list.index(c_issue_account_list[i])
            # 更新表c的I-L列
            c_ws['I' + str(i+3)] = a_id_cards_list[a_index]
            c_ws['J' + str(i+3)] = a_issue_account_list[a_index]
            c_ws['k' + str(i+3)] = a_name_list[a_index]
            c_ws['L' + str(i+3)] = a_project_list[a_index]
            # 如果C表的账号在B表三列中存在
            if row_num >= 0:
                all_account = [b_account_list[row_num], b_card_list[row_num], b_old_account_list[row_num]]
                # 如果A表账户在B表三个值中存在
                if a_issue_account_list[a_index] in all_account:
                    # C表H列，“一套”
                    c_ws['H' + str(i+3)] = "一套"
                    # C表N列：“OK”
                    c_ws['N' + str(i+3)] = "OK"
                else:
                    # 非一套，更新表c代发账户，M列为C表的原代发账户
                    # C表H列，"非一套，已将原代发账户替换为一卡通账户"
                    c_ws['H' + str(i+3)] = "非一套，已将原代发账户替换为一卡通账户"
                    # C表N列：“OK”
                    c_ws['N' + str(i+3)] = "OK"
                    # C表M列，保留原代发账户数据
                    old_account = c_issue_account_list[i]
                    c_ws['M' + str(i+3)] =  old_account
                    # 替换C表D列的代发账户为A表的一卡通账户
                    c_ws['D' + str(i+3)] = a_issue_account_list[a_index]
                    set_row_bg(yellow_bg_fill, i+3)
            else:
                if c_issue_account_list[i] == a_issue_account_list[a_index]:
                    # C表H列，“一套”
                    c_ws['H' + str(i+3)] = "一套"
                    # C表N列：“OK”
                    c_ws['N' + str(i+3)] = "OK"
                else:
                    # 非一套，更新表c代发账户，更新c表的I-L列，M列为C表的原代发账户'
                    # C表H列，"非一套，已将原代发账户替换为一卡通账户"
                    c_ws['H' + str(i+3)] = "非一套，已将原代发账户替换为一卡通账户"
                    # C表N列：“OK”
                    c_ws['N' + str(i+3)] = "OK"
                    # C表M列，保留原代发账户数据
                    c_ws['M' + str(i+3)] = c_issue_account_list[i]
                    # 替换C表D列的代发账户为A表的一卡通账户
                    c_ws['D' + str(i+3)] = a_issue_account_list[a_index]
                    set_row_bg(yellow_bg_fill, i+3)
        else:
            # 判断C表的代发账号是否在B表的三列（账号、卡号、旧帐号）数据中存在
            row_num = -1
            if c_issue_account_list[i] in b_account_list:
                row_num = b_account_list.index(c_issue_account_list[i])
            if c_issue_account_list[i] in b_card_list:
                row_num = b_card_list.index(c_issue_account_list[i])
            if c_issue_account_list[i] in b_old_account_list:
                row_num = b_old_account_list.index(c_issue_account_list[i])
            # 如果存在
            if row_num >= 0:
                # 判断C表和B表的身份证是否一致
                b_id = b_id_cards_list[row_num]
                c_id = c_id_cards_list[i]
                # 复制b表的身份证
                c_ws['O' + str(i+3)] = b_id
                if not b_id:
                    c_ws['N' + str(i+3)] = "暂未提取到数据"
                    set_row_bg(orange_bg_fill, i+3)
                else:
                    if b_id == c_id:
                        # 一致，C表N列“OK”，O列“取B表身份证号”
                        c_ws['N' + str(i+3)] = "OK"
                    elif b_id[-1].isalpha():
                        if b_id.upper() == c_id:
                            c_ws['N' + str(i+3)] = "代发账户数据异常，建议重新提供代发账户"
                            c_ws['P' + str(i+3)] = "身份证末位为小写x"
                            set_row_bg(red_bg_fill, i+3)
                    else:
                        # C表N列，数据异常；O列“取B表身份证号”
                        c_ws['N' + str(i+3)] = "代发账户数据异常，建议重新提供代发账户"
                        # 不一致，判断是否是18和15位新旧身份证问题
                        if down_to_fifteen(c_id) == b_id:
                            # 身份证未升位，P列“旧身份证未升位”
                            c_ws['P' + str(i+3)] = "旧身份证未升位"
                        else:
                            # 身份证完全不一致；P列“身份证完全不一致”
                            c_ws['P' + str(i+3)] = "身份证号完全不一致"
                        set_row_bg(red_bg_fill, i+3)
            # 如果不存在
            else:
                # C表N列“暂未提取到数据”
                c_ws['N' + str(i+3)] = "暂未提取到数据"
                set_row_bg(orange_bg_fill, i+3)


if __name__ == '__main__':
    print("开始执行！")
    start = time.perf_counter()
    local = time.strftime("%Y%m%d%H%M%S", time.localtime())

    a_file_name = "表A-手工数据母表.xlsx"
    a_sht_name = '母表'
    b_file_name = '表B-后台导出数据.xlsx'
    b_sht_name = 'Sheet1'
    c_file_name = "表C-待处理数据.xlsx"
    c_sht_name = 'Sheet1'
    # 更新后的C表名称
    c_new_file_name = "test.xlsx"

    base_dir = os.path.dirname(os.path.abspath(__file__))
    a_file_path = os.path.join(base_dir, a_file_name)
    b_file_path = os.path.join(base_dir, b_file_name)    
    c_file_path = os.path.join(base_dir, c_file_name)

    c_new_file_name = c_new_file_name.split('.')[0] + local + '.' + c_new_file_name.split('.')[1]
    c_new_file_path = os.path.join(base_dir, c_new_file_name)
    c_wb = load_workbook(c_file_path)
    c_ws = c_wb[c_sht_name]
    c_id_cards_tuple = c_ws['C'][2:]
    c_issue_account_tuple = c_ws['D'][2:]
    c_id_cards_list = []
    c_issue_account_list = []
    for i in c_id_cards_tuple:
        c_id_cards_list.append(i.value)
    for i in c_issue_account_tuple:
        c_issue_account_list.append(i.value)
    print("C表数据获取完成")

    a_table = xlrd.open_workbook(a_file_path)
    a_sheet = a_table.sheet_by_name(a_sht_name)
    a_id_cards_list = a_sheet.col_values(2)[1:]
    a_issue_account_list = a_sheet.col_values(3)[1:]
    a_name_list = a_sheet.col_values(4)[1:]
    a_project_list = a_sheet.col_values(6)[1:]
    print("A表数据获取完成")

    b_table = xlrd.open_workbook(b_file_path)
    b_sheet = b_table.sheet_by_name(b_sht_name)
    b_title_list = b_sheet.row_values(0)
    for i in b_title_list:
        if i == '账号':
            account_idx = b_title_list.index(i)
        if i == '卡号':
            card_idx = b_title_list.index(i)
        if i == '旧账号':
            old_account_idx = b_title_list.index(i)
        if i == '证件号码':
            id_cards_idx = b_title_list.index(i)
    b_account_list = b_sheet.col_values(account_idx)[1:]
    b_card_list = b_sheet.col_values(card_idx)[1:]
    b_old_account_list = b_sheet.col_values(old_account_idx)[1:]
    b_id_cards_list = b_sheet.col_values(id_cards_idx)[1:]
    print("B表数据获取完成")

    update_tableC()
    a_table.release_resources()
    b_table.release_resources()
    del a_table
    del b_table
    # c_wb.save(c_file_path)
    c_wb.save(c_new_file_path)

    end = time.perf_counter()
    print("执行成功！")
    print('更新后的C表保存至: %s'%c_new_file_path)
    print('Running time: %s Seconds' %(end-start))
