import os
import time
import datetime
from regex import I
import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

alignment = Alignment(horizontal='center', vertical='center')

def get_tableA():
    wr_A_sht2_data_lists = [] # A表sheet2中需要写入的二维列表
    del_A_sht1_row_num_lists = []  # A表sheet1中需要删除数据的行列表
    wr_A_sht1_data_lists = []  # 从d表需要写入到a表sheet1的数据
    wr_A_sht3_data_lists = []  # 从d表需要写入到a表sheet3的数据
    
    for i in range(len(d_id_cards_list)):
        print("处理第%s条"%i)
        wr_A_sht1_data_list = []
        wr_A_sht2_data_list = []
        wr_A_sht3_data_list = []
        is_exist = False
        # 判断表D的代发账号是否在表B的三列中存在
        b_index = -1
        if d_issue_account_list[i] in b_account_dic.keys():
            b_index = b_account_dic[d_issue_account_list[i]]
        if d_issue_account_list[i] in b_old_account_dic.keys():
            b_index = b_old_account_dic[d_issue_account_list[i]]
        if d_issue_account_list[i] in b_card_dic.keys():
            b_index = b_card_dic[d_issue_account_list[i]]
        # D表的账号在b表三列中存在
        if b_index >= 0:
            if not b_id_cards_list[b_index]:
                d_ws['L' + str(i+2)] = "空值"
                d_ws['L' + str(i+2)].alignment = alignment
                d_ws['M' + str(i+2)] = "数据异常，不予装表"
                d_ws['M' + str(i+2)].alignment = alignment
            else:
                # D表的身份证和B表一致
                if b_id_cards_list[b_index] == d_id_cards_list[i]:
                    # D表身份证在A表存在
                    if d_id_cards_list[i] in a_id_cards_dic.keys():
                        a_index = a_id_cards_dic[d_id_cards_list[i]]
                        # A表账户在B表三个值中存在
                        if a_issue_account_list[a_index] == b_account_list[b_index] or a_issue_account_list[a_index] == b_card_list[b_index] or a_issue_account_list[a_index] == b_old_account_list[b_index]:
                            # D表M列，“已存在，无需装表”
                            d_ws['M' + str(i+2)] = "已存在，无需装表"
                            d_ws['M' + str(i+2)].alignment = alignment
                        else:
                            if b_card_list[b_index].startswith('623055'):
                                d_ws['L' + str(i+2)] = b_id_cards_list[b_index]
                                d_ws['L' + str(i+2)].alignment = alignment
                                d_ws['M' + str(i+2)] = "市民卡替换装表"
                                d_ws['M' + str(i+2)].alignment = alignment
                                a_row_data = a_sheet1[a_index + 2]
                                # 根据身份证判断该数据是否在数据列表里已经存在，避免写入多条数据
                                is_exist = False
                                for item in wr_A_sht2_data_lists:
                                    if item[2] == a_row_data[2].value:
                                        is_exist = True
                                        break
                                # 如果不存在，更新wr_A_sht2_data_lists
                                if not is_exist:
                                    del_A_sht1_row_num_lists.append(a_index+2)
                                    wr_A_sht2_data_list = [a_row_data[j].value for j in range(len(a_row_data))]
                                    wr_A_sht2_data_list.append(datetime.date(year,month,day))
                                    wr_A_sht2_data_list.append(b_status_list[b_index])
                                    wr_A_sht2_data_lists.append(wr_A_sht2_data_list)
                                is_exist = False
                                for item in wr_A_sht1_data_lists:
                                    if item[2] == a_row_data[2].value:
                                        is_exist = True
                                        break
                                # 如果不存在，更新wr_A_sht1_data_lists
                                if not is_exist:
                                    wr_A_sht1_data_list.append(d_ws['A' + str(i+2)].value)
                                    wr_A_sht1_data_list.append(d_ws['J' + str(i+2)].value)
                                    wr_A_sht1_data_list.append(d_ws['G' + str(i+2)].value)
                                    wr_A_sht1_data_list.append(b_account_list[b_index])
                                    wr_A_sht1_data_list.append(d_ws['C' + str(i+2)].value)
                                    wr_A_sht1_data_list.append(d_ws['E' + str(i+2)].value)
                                    wr_A_sht1_data_list.append(d_ws['K' + str(i+2)].value)
                                    wr_A_sht1_data_list.append(datetime.date(year,month,day))
                                    wr_A_sht1_data_lists.append(wr_A_sht1_data_list)
                            else:
                                d_ws['M' + str(i+2)] = "未按一卡通账号代发，不予装表"
                                d_ws['M' + str(i+2)].alignment = alignment
                    # D表身份证在A表不存在
                    else:
                        d_ws['L' + str(i+2)] = b_id_cards_list[b_index]
                        d_ws['L' + str(i+2)].alignment = alignment
                        d_ws['M' + str(i+2)] = "新装表"
                        d_ws['M' + str(i+2)].alignment = alignment
                        is_exist = False
                        for item in wr_A_sht1_data_lists:
                            if item[2] == d_id_cards_list[i]:
                                is_exist = True
                                break
                        if not is_exist: 
                            # 表D内容写入表A
                            wr_A_sht1_data_list.append(d_ws['A' + str(i+2)].value)
                            wr_A_sht1_data_list.append(d_ws['J' + str(i+2)].value)
                            wr_A_sht1_data_list.append(d_ws['G' + str(i+2)].value)
                            # B表账号需写入A表
                            wr_A_sht1_data_list.append(b_account_list[b_index])
                            # 表D内容写入表A
                            wr_A_sht1_data_list.append(d_ws['C' + str(i+2)].value)
                            wr_A_sht1_data_list.append(d_ws['E' + str(i+2)].value)
                            wr_A_sht1_data_list.append(d_ws['K' + str(i+2)].value)
                            # 装表日期
                            wr_A_sht1_data_list.append(datetime.date(year,month,day))
                            wr_A_sht1_data_lists.append(wr_A_sht1_data_list)
                # D表的身份证和B表不一致
                else:
                    # D表身份证在A表存在
                    if d_id_cards_list[i] in a_id_cards_dic.keys():
                        d_ws['M' + str(i+2)] = "未按一卡通账号代发，不予装表"
                        d_ws['M' + str(i+2)].alignment = alignment
                    # D表身份证在A表不存在
                    else:
                        d_ws['L' + str(i+2)] = b_id_cards_list[b_index]
                        d_ws['L' + str(i+2)].alignment = alignment
                        d_ws['M' + str(i+2)] = "数据异常，不予装表"
                        d_ws['M' + str(i+2)].alignment = alignment
        # D表的账号在B表不存在
        else:
            # D表身份证在A表存在
            if d_id_cards_list[i] in a_id_cards_dic.keys():
                d_ws['M' + str(i+2)] = "未按一卡通账号代发，不予装表"
                d_ws['M' + str(i+2)].alignment = alignment
            else:
                d_ws['M' + str(i+2)] = "暂未提取到数据，不予装表"
                d_ws['M' + str(i+2)].alignment = alignment
                # 根据身份证判断是否在wr_A_sht3_data_lists中存在，不存在就更新
                is_exist = False
                for item in wr_A_sht3_data_lists:
                    if item[2] == d_id_cards_list[i]:
                        is_exist = True
                        break
                if not is_exist:
                    wr_A_sht3_data_list.append(d_ws['A' + str(i+2)].value)
                    wr_A_sht3_data_list.append(d_ws['J' + str(i+2)].value)
                    wr_A_sht3_data_list.append(d_ws['G' + str(i+2)].value)
                    wr_A_sht3_data_list.append(d_ws['B' + str(i+2)].value)
                    wr_A_sht3_data_list.append(d_ws['C' + str(i+2)].value)
                    wr_A_sht3_data_list.append(d_ws['E' + str(i+2)].value)
                    wr_A_sht3_data_list.append(d_ws['K' + str(i+2)].value)
                    wr_A_sht3_data_lists.append(wr_A_sht3_data_list)

    return wr_A_sht2_data_lists, del_A_sht1_row_num_lists, wr_A_sht1_data_lists, wr_A_sht3_data_lists

def update_tableA(wr_A_sht2_data_lists, del_A_sht1_row_num_lists, wr_A_sht1_data_lists, wr_A_sht3_data_lists):
    wr_sht2_len = len(wr_A_sht2_data_lists)
    del_sht1_len = len(del_A_sht1_row_num_lists)
    wr_sht1_len = len(wr_A_sht1_data_lists)
    wr_sht3_len = len(wr_A_sht3_data_lists)
    
    # 写入A表Sheet2
    if wr_sht2_len:
        print("正在写入A表Sheet2 ...")
        for r in range(wr_sht2_len):
            for c in range(len(wr_A_sht2_data_lists[0])):
                a_sheet2.cell(a_sht2_max_row + r + 1, c + 1).value = wr_A_sht2_data_lists[r][c]
                a_sheet2['H' + str(a_sht2_max_row + r + 1)].number_format = 'yyyy/m/d'                
                a_sheet2['I' + str(a_sht2_max_row + r + 1)].number_format = 'yyyy/m/d'
                a_sheet2.cell(a_sht2_max_row + r + 1, c + 1).alignment = alignment
    # 删除A表sheet1
    if del_sht1_len:
        print("正在删除A表Sheet1 ...")
        for row_num in range(del_sht1_len,0,-1):
            a_sheet1.delete_rows(del_A_sht1_row_num_lists[row_num-1])
    # 写入A表Sheet1
    if wr_sht1_len:
        print("正在写入A表Sheet1 ...")
        a_sht1_max_row = a_sheet1.max_row
        for r in range(wr_sht1_len):
            for c in range(len(wr_A_sht1_data_lists[0])):
                a_sheet1.cell(a_sht1_max_row + r + 1, c + 1).value = wr_A_sht1_data_lists[r][c]
                a_sheet1['H' + str(a_sht1_max_row + r + 1)].number_format = 'yyyy/m/d'
                a_sheet1.cell(a_sht1_max_row + r + 1, c + 1).alignment = alignment
    # 写入A表Sheet3
    if wr_sht3_len:
        print("正在写入A表Sheet3 ...")
        for r in range(wr_sht3_len):
            for c in range(len(wr_A_sht3_data_lists[0])):
                a_sheet3.cell(a_sht3_max_row + r + 1, c + 1).value = wr_A_sht3_data_lists[r][c]
                a_sheet3.cell(a_sht3_max_row + r + 1, c + 1).alignment = alignment

if __name__ == '__main__':
    print("开始执行！")
    start = time.perf_counter()
    local = time.strftime("%Y%m%d%H%M%S", time.localtime())
    date = time.localtime()
    year, month, day = date.tm_year, date.tm_mon, date.tm_mday

    # A\B\D表名和sheet名
    a_file_name = "表A-手工数据母表.xlsx"
    a_sht1_name = '母表'
    a_sht2_name = '已删除账号信息表'
    a_sht3_name = '待发省联社取数'

    b_file_name = '表B-后台导出数据.xlsx'
    b_sht_name = 'Sheet1'

    d_file_name = "表D-结果数据.xlsx"
    d_sht_name = 'Sheet1'

    # 更新后的A表\D表名称
    a_new_file_name = '更新后的表A.xlsx'
    d_new_file_name = "更新后的表D.xlsx"
    # A\B\D文件路径
    base_dir = os.path.dirname(os.path.abspath(__file__))
    a_file_path = os.path.join(base_dir, a_file_name)
    b_file_path = os.path.join(base_dir, b_file_name)    
    d_file_path = os.path.join(base_dir, d_file_name)
    
    # 提取D表身份证、账号数据
    d_wb = load_workbook(d_file_path)
    d_ws = d_wb[d_sht_name]
    d_id_cards_tuple = d_ws['G'][1:]
    d_issue_account_tuple = d_ws['B'][1:]
    d_id_cards_list = []
    d_issue_account_list = []
    for i in d_id_cards_tuple:
        d_id_cards_list.append(i.value)
    for i in d_issue_account_tuple:
        d_issue_account_list.append(i.value)
    print("D表数据获取完成")

    # 提取A表身份证、账号数据
    a_table = load_workbook(a_file_path)
    a_sheet1 = a_table[a_sht1_name]
    a_sheet2 = a_table[a_sht2_name]
    a_sheet3 = a_table[a_sht3_name]

    a_id_cards_tuple = a_sheet1['C'][1:]
    a_issue_account_tuple = a_sheet1['D'][1:]
    a_id_cards_list = []
    a_issue_account_list = []
    for i in a_id_cards_tuple:
        a_id_cards_list.append(i.value)
    for i in a_issue_account_tuple:
        a_issue_account_list.append(i.value)

    a_idx = [x for x in range(len(a_id_cards_list))]
    a_id_cards_dic = dict(zip(a_id_cards_list, a_idx))
    
    # A表各sheet最大行数
    a_sht1_max_row = a_sheet1.max_row
    a_sht2_max_row = a_sheet2.max_row
    a_sht3_max_row = a_sheet3.max_row
    print("A表数据获取完成")

    # B表列名、账号、卡号、旧帐号、身份证号、状态数据
    b_table = xlrd.open_workbook(b_file_path)
    b_sheet = b_table.sheet_by_name(b_sht_name)
    b_title_list = b_sheet.row_values(0)
    for i in b_title_list:
        if i == '账号':
            account_idx = b_title_list.index(i)
        if i == '卡号':
            card_idx = b_title_list.index(i)
        if i == '旧账号':
            old_account_idx = b_title_list.index(i)
        if i == '证件号码':
            id_cards_idx = b_title_list.index(i)
        if i == "当前状态":
            status_idx = b_title_list.index(i)
    b_account_list = b_sheet.col_values(account_idx)[1:]
    b_card_list = b_sheet.col_values(card_idx)[1:]
    b_old_account_list = b_sheet.col_values(old_account_idx)[1:]
    b_id_cards_list = b_sheet.col_values(id_cards_idx)[1:]
    b_status_list = b_sheet.col_values(status_idx)[1:]
    
    b_idx = [x for x in range(len(b_account_list))]
    b_account_dic = dict(zip(b_account_list, b_idx))
    b_card_dic = dict(zip(b_card_list, b_idx))
    b_old_account_dic = dict(zip(b_old_account_list, b_idx))
    print("B表数据获取完成")

    wr_A_sht2_data_lists, del_A_sht1_row_num_lists, wr_A_sht1_data_lists, wr_A_sht3_data_lists = get_tableA()
    # print(wr_A_sht2_data_lists)
    # print(del_A_sht1_row_num_lists)
    # print(wr_A_sht1_data_lists)
    # print(wr_A_sht3_data_lists)
    update_tableA(wr_A_sht2_data_lists, del_A_sht1_row_num_lists, wr_A_sht1_data_lists, wr_A_sht3_data_lists)

    # 关闭B表
    b_table.release_resources()
    del b_table

    # 保存更新后的A表和D表
    a_new_file_name = a_new_file_name.split('.')[0] + local + '.' + a_new_file_name.split('.')[1]
    a_new_file_path = os.path.join(base_dir, a_new_file_name)
    a_table.save(a_new_file_path)
    d_new_file_name = d_new_file_name.split('.')[0] + local + '.' + d_new_file_name.split('.')[1]
    d_new_file_path = os.path.join(base_dir, d_new_file_name)
    d_wb.save(d_new_file_path)

    end = time.perf_counter()
    print("执行成功！")
    print('更新后的A表保存至: %s'%a_new_file_path)
    print('更新后的D表保存至: %s'%d_new_file_path)
    print('Running time: %s Seconds' %(end-start))
