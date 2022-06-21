import os
import time
import datetime
import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

alignment = Alignment(horizontal='center', vertical='center')

def get_tableA():
    row_data_lists= [] # A表sheet1中需要删除的二维列表
    row_num_list = []  # A表sheet1中需要删除数据的行列表
    # 判断A表账号是否在B表中存在
    for i in range(len(a_issue_account_list)):
        print("处理第%s条"%i)
        a_issue_account = a_issue_account_list[i]
        b_index = -1
        if a_issue_account in b_account_dic.keys():
            b_index = b_account_dic[a_issue_account]
        if a_issue_account in b_card_dic.keys():
            b_index = b_card_dic[a_issue_account]
        if a_issue_account in b_old_account_dic.keys():
            b_index = b_old_account_dic[a_issue_account]
        if b_index == -1:
            continue
        if b_status_list[b_index] == "关户":
            a_row_data = a_sheet1[i + 2]  # 该行数据
            row_num_list.append(i+2)
            row_data_list = [a_row_data[j].value for j in range(len(a_row_data))]
            row_data_lists.append(row_data_list)
    return row_data_lists, row_num_list

def update_tableA(row_data_lists, row_num_list):  
    # A表需要更新
    if row_data_lists and row_num_list:
        # 写入A表Sheet2
        for r in range(len(row_data_lists)):
            for c in range(len(row_data_lists[0])):
                a_sheet2.cell(a_sht2_max_row + r + 1, c + 1).value = row_data_lists[r][c]
                a_sheet2['H' + str(a_sht2_max_row + r + 1)].number_format = 'yyyy/m/d'
                a_sheet2.cell(a_sht2_max_row + r + 1, c + 1).alignment = alignment
                a_sheet2['I' + str(a_sht2_max_row + r + 1)] = datetime.date(year,month,day)
                a_sheet2['I' + str(a_sht2_max_row + r + 1)].number_format = 'yyyy/m/d'
                a_sheet2['I' + str(a_sht2_max_row + r + 1)].alignment = alignment
                a_sheet2['J' + str(a_sht2_max_row + r + 1)] = '关户'
                a_sheet2['J' + str(a_sht2_max_row + r + 1)].alignment = alignment
        # 删除A表数据
        for row_num in range(len(row_num_list),0,-1):
            a_sheet1.delete_rows(row_num_list[row_num-1])
    else:
        print("A表没有需要更新的数据")
        return


if __name__ == '__main__':
    print("开始执行！")
    start = time.perf_counter()
    local = time.strftime("%Y%m%d%H%M%S", time.localtime())
    date = time.localtime()
    year, month, day = date.tm_year, date.tm_mon, date.tm_mday


    a_file_name = "表A-手工数据母表.xlsx"
    a_sht1_name = '母表'
    a_sht2_name = '已删除账号信息表'

    b_file_name = '表B-后台导出数据.xlsx'
    b_sht_name = 'Sheet1'

    # 更新后的A表名称
    a_new_file_name = '更新后的表A.xlsx'

    base_dir = os.path.dirname(os.path.abspath(__file__))
    a_file_path = os.path.join(base_dir, a_file_name)
    b_file_path = os.path.join(base_dir, b_file_name)

    a_table = load_workbook(a_file_path)
    a_sheet1 = a_table[a_sht1_name]
    a_sheet2 = a_table[a_sht2_name]

    a_id_cards_tuple = a_sheet1['C'][1:]
    a_issue_account_tuple = a_sheet1['D'][1:]
    a_issue_account_list = []
    for i in a_issue_account_tuple:
        a_issue_account_list.append(i.value)

    a_sht2_max_row = a_sheet2.max_row
    print("A表数据获取完成")

    b_table = xlrd.open_workbook(b_file_path)
    b_sheet = b_table.sheet_by_name(b_sht_name)
    b_title_list = b_sheet.row_values(0)
    for i in b_title_list:
        if i == '账号':
            account_idx = b_title_list.index(i)
        if i == '卡号':
            card_idx = b_title_list.index(i)
        if i == '旧账号':
            old_account_idx = b_title_list.index(i)
        if i == "当前状态":
            status_idx = b_title_list.index(i)
    b_account_list = b_sheet.col_values(account_idx)[1:]
    b_card_list = b_sheet.col_values(card_idx)[1:]
    b_old_account_list = b_sheet.col_values(old_account_idx)[1:]
    b_status_list = b_sheet.col_values(status_idx)[1:]
    
    b_idx = [x for x in range(len(b_account_list))]
    b_account_dic = dict(zip(b_account_list, b_idx))
    b_card_dic = dict(zip(b_card_list, b_idx))
    b_old_account_dic = dict(zip(b_old_account_list, b_idx))
    print("B表数据获取完成")

    row_data_lists, row_num_list = get_tableA()
    update_tableA(row_data_lists, row_num_list)

    b_table.release_resources()
    del b_table


    # 保存更新后的A表
    a_new_file_name = a_new_file_name.split('.')[0] + local + '.' + a_new_file_name.split('.')[1]
    a_new_file_path = os.path.join(base_dir, a_new_file_name)
    a_table.save(a_new_file_path)


    end = time.perf_counter()
    print("执行成功！")
    print('更新后的A表保存至: %s'%a_new_file_path)
    print('Running time: %s Seconds' %(end-start))
