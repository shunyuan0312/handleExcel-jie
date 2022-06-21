import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, fills, Alignment
import time
import datetime
import os
from openpyxl.formatting import Rule
base_dir = os.path.dirname(os.path.abspath(__file__))
# a_file_name = "表A-手工数据母表.xlsx"   
# a_file_path = os.path.join(base_dir, a_file_name)
# data = xlrd.open_workbook(a_file_path)
# table = data.sheets()[0]
# list1 = table.row_values(1)
# print(list1)
#[1.0, '测试单位1', '42272319591018125X', '6210137281175123', '杨明想', 1.0, '测试项目1', 44562.0]


c_file_name = "test.xlsx"
c_sht_name = 'Sheet1'
# c_sht_name = '母表'
c_file_path = os.path.join(base_dir, c_file_name)
c_wb = load_workbook(c_file_path)
c_ws = c_wb[c_sht_name]
c_ws2 = c_wb['Sheet2']
# c_ws2['I' + str(3)] = time.strftime("%Y/%m/%d", time.localtime())
# c_ws2['I' + str(3)] = datetime.date(2022,4,13)
# c_ws2['I' + str(3)].number_format = 'yyyy/m/d'
# c_ws2['H' + str(3)] = '123'
# max_row = c_ws2.max_row
# print('1', max_row)
# c_ws2.delete_rows(3)
# max_row = c_ws2.max_row
# print('2', max_row)
range_cell = 'A1:A10'
rule = Rule(type='cellIs', formula=["=MAX("+ range_cell +")"])
c_ws.conditional_formatting.add(range_cell, rule)

# c_ws.delete_rows(30)
# c_tuple = c_ws[2]
for i in range(len(c_tuple)):
    cell = c_tuple[i].value
    letter = chr(i + 65)
    c_ws2[letter+str(max_row+1)] = cell
    


# for i in 16:
#     c_ws['A5'].fill = red
# c_ws.row_dimensions[7].height = 100
# c_ws.row_dimensions[7].fill = PatternFill(fill_type="solid",fgColor="ff0000")
c_wb.save('表D-结果数据1.xlsx')